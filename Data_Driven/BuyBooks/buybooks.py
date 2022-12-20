# Generated by Selenium IDE
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import Select

from selenium.common.exceptions import NoSuchElementException

class FileExcelReader:
    file = ""
    sheetName = ""

    def __init__(self, file, sheetName):
        self.file = file
        self.sheetName = sheetName

    def getRowCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_row)

    def getColumnCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_column)

    def readData(self, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    def writeData(self, data, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        wordbook.save(self.file)

class TestBuyBooks():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.maximize_window()
    self.vars = {}
    self.driver.get("https://mybk.hcmut.edu.vn/my/logoutSSO.action")
    self.driver.get('https://sso.hcmut.edu.vn/cas/login?service=https://mybk.hcmut.edu.vn/my/homeSSO.action')
    username = self.driver.find_element(By.NAME,"username")
    password = self.driver.find_element(By.NAME,"password")  
    submitBtn = self.driver.find_element(By.NAME,"submit")  
    
    username.send_keys("kien.ha04102002")
    password.send_keys("kien05111")
    submitBtn.click()
  
  def teardown_method(self):
    self.driver.get("https://mybk.hcmut.edu.vn/my/homeSSO.action")
    logoutBtn = self.driver.find_element(By.XPATH,"/html/body/div[2]/div/div/div[2]/div/a[1]")
    logoutBtn.click()
    self.driver.quit()
  
  
  def test_buybooks(self,quantity,flow,address,expectedResult ):
    self.driver.get('https://mybk.hcmut.edu.vn/bookstore/')

    if flow == "Normal":
        self.driver.get("https://mybk.hcmut.edu.vn/bookstore/product/2")
        quantityInput = self.driver.find_element(By.XPATH,'//*[@id="txtQuantity"]')
        quantityInput.clear()
        quantityInput.send_keys(quantity)
        time.sleep(2)
        addToCartBtn = self.driver.find_element(By.ID,'bntAddToCart')
        addToCartBtn.click()
        time.sleep(2)
        self.driver.get("https://mybk.hcmut.edu.vn/bookstore/cart")
        time.sleep(2)
        while self.driver.current_url != 'https://mybk.hcmut.edu.vn/bookstore/checkout':
            self.driver.get("https://mybk.hcmut.edu.vn/bookstore/cart")
            time.sleep(2)
            checkOutBtn = self.driver.find_element(By.XPATH,'//*[@id="bntCheckout"]')
            checkOutBtn.click()
            time.sleep(2) 
        informationBtn = self.driver.find_element(By.XPATH,'//*[@id="headingTwo"]/h4/a')
        informationBtn.click()
        time.sleep(2)
        addressInput = self.driver.find_element(By.ID,"txtAddress")
        addressInput.clear()
        addressInput.send_keys(address)
        time.sleep(2)
        assert self.driver.current_url == 'https://mybk.hcmut.edu.vn/bookstore/checkout'
    elif flow =="Alter1":
        self.driver.get("https://mybk.hcmut.edu.vn/bookstore/product/2")
        quantityInput = self.driver.find_element(By.XPATH,'//*[@id="txtQuantity"]')
        quantityInput.clear()
        quantityInput.send_keys(quantity)
        time.sleep(2)
        addToCartBtn = self.driver.find_element(By.ID,'bntAddToCart')
        addToCartBtn.click()
        time.sleep(2)
        self.driver.get("https://mybk.hcmut.edu.vn/bookstore/cart")
        time.sleep(2)
        self.driver.get("https://mybk.hcmut.edu.vn/bookstore/product/2")
        quantityInput = self.driver.find_element(By.XPATH,'//*[@id="txtQuantity"]')
        quantityInput.clear()
        quantityInput.send_keys(quantity)
        time.sleep(2)
        addToCartBtn = self.driver.find_element(By.ID,'bntAddToCart')
        addToCartBtn.click()
        time.sleep(2)
        self.driver.get("https://mybk.hcmut.edu.vn/bookstore/cart")
        time.sleep(2)
        while self.driver.current_url != 'https://mybk.hcmut.edu.vn/bookstore/checkout':
            self.driver.get("https://mybk.hcmut.edu.vn/bookstore/cart")
            time.sleep(2)
            checkOutBtn = self.driver.find_element(By.XPATH,'//*[@id="bntCheckout"]')
            checkOutBtn.click()
            time.sleep(2) 
        informationBtn = self.driver.find_element(By.XPATH,'//*[@id="headingTwo"]/h4/a')
        informationBtn.click()
        time.sleep(2)
        addressInput = self.driver.find_element(By.ID,"txtAddress")
        addressInput.clear()
        addressInput.send_keys(address)
        time.sleep(2)
        assert self.driver.current_url == 'https://mybk.hcmut.edu.vn/bookstore/checkout'

    elif flow == "Exception":
        self.driver.get("https://mybk.hcmut.edu.vn/bookstore/product/2")
        quantityInput = self.driver.find_element(By.XPATH,'//*[@id="txtQuantity"]')
        quantityInput.clear()
        quantityInput.send_keys(quantity)
        time.sleep(2)
        addToCartBtn = self.driver.find_element(By.ID,'bntAddToCart')
        addToCartBtn.click()
        time.sleep(2)
        self.driver.get("https://mybk.hcmut.edu.vn/bookstore/cart")
        time.sleep(2)
        while self.driver.current_url != 'https://mybk.hcmut.edu.vn/bookstore/checkout':
            self.driver.get("https://mybk.hcmut.edu.vn/bookstore/cart")
            time.sleep(2)
            checkOutBtn = self.driver.find_element(By.XPATH,'//*[@id="bntCheckout"]')
            checkOutBtn.click()
            time.sleep(2) 
        informationBtn = self.driver.find_element(By.XPATH,'//*[@id="headingTwo"]/h4/a')
        informationBtn.click()
        time.sleep(2)
        time.sleep(2)
        addressInput = self.driver.find_element(By.ID,"txtAddress")
        addressInput.clear()
        time.sleep(1)
        detailBtn = self.driver.find_element(By.XPATH,'//*[@id="headingFive"]/h4/a')
        detailBtn .click()
        time.sleep(1)
        confirmBtn = self.driver.find_element(By.XPATH,'//*[@id="bntOrder"]/span')
        confirmBtn.click()
        time.sleep(2)
        assert self.driver.find_element(By.XPATH,'//*[@id="swal2-content"]')       

    else:
        self.driver.get("https://mybk.hcmut.edu.vn/bookstore/product/2")
        quantityInput = self.driver.find_element(By.XPATH,'//*[@id="txtQuantity"]')
        quantityInput.clear()
        quantityInput.send_keys(quantity)
        time.sleep(2)
        addToCartBtn = self.driver.find_element(By.ID,'bntAddToCart')
        addToCartBtn.click()
        time.sleep(3)
        if expectedResult == "InvalidQuantity":
            assert self.driver.find_element(By.XPATH,'//*[@id="swal2-content"]')
        #     notification = self.driver.find_element(By.XPATH,'//*[@id="swal2-content"]')
        #     # assert notification == 'Số lượng mua phải lớn hơn 0' or  notification == 'Số lượng mua tối đa 5 cuốn cho 1 tên sách'
        #     assert notification != ""
        #     return True
        # else:
        #     return True


  
if __name__ == "__main__":
    excel = FileExcelReader('SecB_buybook_data.xlsx', 'Sheet1')

    test = TestBuyBooks()
    test.setup_method()
    nRows = excel.getRowCount()
    for row in range(2, nRows + 1):
        quantity = excel.readData(row,1)
        flow = excel.readData(row,2)
        address = excel.readData(row,3)
        expectedResult = excel.readData(row,4)

        if address is None:
            address = ""

            
        try:
            result = test.test_buybooks(quantity,flow,address,expectedResult)
            excel.writeData("Passed",row,5)
        except:
            excel.writeData("Failed",row,5)


    test.teardown_method()